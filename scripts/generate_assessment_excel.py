import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

def parse_md_assignments(filepath):
    """
    Parses markdown assignment files for students and their assigned tasks.
    Handles escaped pipe characters inside table cells.
    """
    students = {}
    if not os.path.exists(filepath):
        print(f"Warning: File {filepath} tidak ditemukan.")
        return students

    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            l = line.strip()
            if l.startswith('|') and not l.startswith('| Task') and not l.startswith('| :---'):
                # Replace escaped pipe characters \| to prevent incorrect column splitting
                l_clean = l.replace(r'\|', '---PIPE---')
                parts = [p.strip().replace('---PIPE---', '|') for p in l_clean.split('|')]
                if len(parts) >= 6:
                    task_desc = parts[1]
                    path = parts[2]
                    method = parts[3]
                    student_name = parts[4]
                    if student_name and student_name != '-' and not student_name.startswith('---'):
                        if student_name not in students:
                            students[student_name] = []
                        students[student_name].append({
                            'task': task_desc,
                            'path': path,
                            'method': method
                        })
    return students

def create_assessment_workbook():
    wb = Workbook()
    
    # -------------------------------------------------------------
    # Define Design System & Typography
    # -------------------------------------------------------------
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    subtitle_font = Font(name=font_family, size=11, italic=True, color="595959")
    section_font = Font(name=font_family, size=13, bold=True, color="1F4E78")
    
    # Header styling: Navy Blue (#1F4E78) background with bold white text
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    legend_header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    
    data_font = Font(name=font_family, size=10)
    bold_data_font = Font(name=font_family, size=10, bold=True)
    
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    header_border_side = Side(border_style="thin", color="1B365D")
    header_border = Border(left=header_border_side, right=header_border_side, top=header_border_side, bottom=header_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # -------------------------------------------------------------
    # Sheet 1: Petunjuk & Legenda
    # -------------------------------------------------------------
    ws_legenda = wb.active
    ws_legenda.title = "Petunjuk & Legenda"
    ws_legenda.views.sheetView[0].showGridLines = True
    
    ws_legenda['B2'] = "PETUNJUK & LEGENDA PENILAIAN MAHASISWA RPL 2026"
    ws_legenda['B2'].font = title_font
    
    ws_legenda['B3'] = "Panduan Aturan Gradasi Nilai Akhir & Status Kelulusan Sistem Penilaian"
    ws_legenda['B3'].font = subtitle_font
    
    # Aturan Gradasi Nilai Table Header
    ws_legenda['B5'] = "1. ATURAN GRADASI NILAI & STATUS"
    ws_legenda['B5'].font = section_font
    
    legenda_headers = ["Jumlah Task Selesai", "Warning Merah", "Nilai Akhir", "Status", "Keterangan Kualifikasi"]
    for col_idx, h in enumerate(legenda_headers, start=2):
        cell = ws_legenda.cell(row=6, column=col_idx, value=h)
        cell.fill = legend_header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = header_border
    ws_legenda.row_dimensions[6].height = 25
    
    legenda_data = [
        [4, "T", "A", "LULUS", "Sangat Baik: Menyelesaikan seluruh (4) task penugasan."],
        [3, "T", "B+", "LULUS", "Baik: Menyelesaikan 3 task penugasan."],
        [2, "T", "B", "LULUS", "Cukup: Menyelesaikan 2 task penugasan."],
        [1, "T", "D", "BELUM LULUS", "Kurang: Hanya menyelesaikan 1 task penugasan."],
        [0, "T", "E", "BELUM LULUS", "Sangat Kurang: Tidak menyelesaikan task penugasan."],
        ["Bebas (0-4)", "Y", "E", "BELUM LULUS", "Diskualifikasi: Mendapatkan Warning Merah (Pelanggaran/Terindikasi Bermasalah)."]
    ]
    
    for r_idx, row_data in enumerate(legenda_data, start=7):
        ws_legenda.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(row_data, start=2):
            cell = ws_legenda.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c_idx in [2, 3, 4, 5]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
            
            # Custom font colors for Nilai & Status columns
            if c_idx == 4: # Nilai Akhir
                if val == "A":
                    cell.font = Font(name=font_family, size=10, bold=True, color="385723")
                elif val in ["B+", "B"]:
                    cell.font = Font(name=font_family, size=10, bold=True, color="1B365D")
                elif val in ["D", "E"]:
                    cell.font = Font(name=font_family, size=10, bold=True, color="C00000")
            elif c_idx == 5: # Status
                if val == "LULUS":
                    cell.font = Font(name=font_family, size=10, bold=True, color="385723")
                else:
                    cell.font = Font(name=font_family, size=10, bold=True, color="C00000")

    # Section 2: Instructions
    ws_legenda['B15'] = "2. PETUNJUK PENGGUNAAN SPREADSHEET"
    ws_legenda['B15'].font = section_font
    
    instructions = [
        "a. Pilih sheet 'Penilaian Kelas B' atau 'Penilaian Kelas G' sesuai kelas yang ingin dinilai.",
        "b. Status Pengerjaan Task (Task 1 s.d Task 4) diisi dengan memilih nilai dari dropdown list ('Belum' / 'Selesai'). Default: 'Belum'.",
        "c. Hover / Arahkan kursor ke cell 'Task 1' s.d 'Task 4' untuk melihat detail deskripsi tugas dan file path masing-masing mahasiswa.",
        "d. Kolom 'Warning Merah' digunakan untuk memberikan sanksi/warning (dropdown 'Y' / 'T'). Default: 'T'.",
        "e. Kolom 'Total Task Selesai', 'Nilai Akhir', dan 'Status' terhitung secara OTOMATIS menggunakan rumus Excel:",
        "    - Total Task Selesai = COUNTIF(D5:G5, \"Selesai\")",
        "    - Nilai Akhir = IF(H5=\"Y\",\"E\",IF(I5=4,\"A\",IF(I5=3,\"B+\",IF(I5=2,\"B\",IF(I5=1,\"D\",\"E\")))))",
        "    - Status = IF(J5=\"E\",\"BELUM LULUS\",\"LULUS\")",
        "f. Jangan mengubah atau menghapus rumus pada kolom Total Task Selesai, Nilai Akhir, dan Status."
    ]
    
    for i, inst in enumerate(instructions, start=16):
        cell = ws_legenda.cell(row=i, column=2, value=inst)
        cell.font = data_font
        ws_legenda.row_dimensions[i].height = 18

    ws_legenda.column_dimensions['A'].width = 3
    ws_legenda.column_dimensions['B'].width = 22
    ws_legenda.column_dimensions['C'].width = 16
    ws_legenda.column_dimensions['D'].width = 14
    ws_legenda.column_dimensions['E'].width = 16
    ws_legenda.column_dimensions['F'].width = 80

    # -------------------------------------------------------------
    # Parse Classes Data
    # -------------------------------------------------------------
    class_b_data = parse_md_assignments('docs/guidence rpl 2026/Penugasan_RPL_Kelas_B.md')
    class_g_data = parse_md_assignments('docs/guidence rpl 2026/Penugasan_RPL_Kelas_G.md')
    
    def populate_class_sheet(ws, class_name, subtitle, students_dict):
        ws.views.sheetView[0].showGridLines = True
        
        # Title Block
        ws.cell(row=1, column=1, value=f"DAFTAR PENILAIAN MAHASISWA RPL 2026 - {class_name.upper()}").font = title_font
        ws.cell(row=2, column=1, value=subtitle).font = subtitle_font
        
        # Table Headers
        headers = [
            "No", "NIM", "Nama Mahasiswa",
            "Task 1", "Task 2", "Task 3", "Task 4",
            "Warning Merah", "Total Task Selesai", "Nilai Akhir", "Status"
        ]
        
        header_row = 4
        ws.row_dimensions[header_row].height = 28
        
        for col_idx, text in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = header_border
            
        start_row = 5
        current_row = start_row
        
        for idx, (student_name, tasks) in enumerate(students_dict.items(), start=1):
            ws.row_dimensions[current_row].height = 22
            is_even = (idx % 2 == 0)
            row_fill = zebra_fill if is_even else PatternFill(fill_type=None)
            
            # Col A: No
            cell_no = ws.cell(row=current_row, column=1, value=idx)
            cell_no.alignment = align_center
            cell_no.font = data_font
            cell_no.border = thin_border
            if is_even: cell_no.fill = row_fill
            
            # Col B: NIM (Default empty string)
            cell_nim = ws.cell(row=current_row, column=2, value="")
            cell_nim.alignment = align_center
            cell_nim.font = data_font
            cell_nim.border = thin_border
            if is_even: cell_nim.fill = row_fill
            
            # Col C: Nama Mahasiswa
            cell_nama = ws.cell(row=current_row, column=3, value=student_name)
            cell_nama.alignment = align_left
            cell_nama.font = bold_data_font
            cell_nama.border = thin_border
            if is_even: cell_nama.fill = row_fill
            
            # Col D-G: Task 1 - Task 4 (Default = 'Belum')
            for task_num in range(1, 5):
                col_idx = 3 + task_num
                cell_task = ws.cell(row=current_row, column=col_idx, value="Belum")
                cell_task.alignment = align_center
                cell_task.font = data_font
                cell_task.border = thin_border
                if is_even: cell_task.fill = row_fill
                
                # Add comment with task details if available
                if task_num <= len(tasks):
                    t_info = tasks[task_num - 1]
                    comment_text = f"Task {task_num}:\n{t_info['task']}\n\nPath:\n{t_info['path']}"
                    if t_info['method'] and t_info['method'] != '-':
                        comment_text += f"\n\nMethod:\n{t_info['method']}"
                    cell_task.comment = Comment(comment_text, "Panduan RPL 2026")
            
            # Col H: Warning Merah (Default = 'T')
            cell_wm = ws.cell(row=current_row, column=8, value="T")
            cell_wm.alignment = align_center
            cell_wm.font = data_font
            cell_wm.border = thin_border
            if is_even: cell_wm.fill = row_fill
            
            # Col I: Total Task Selesai
            # Formula: =COUNTIF(D5:G5, "Selesai")
            cell_total = ws.cell(row=current_row, column=9, value=f'=COUNTIF(D{current_row}:G{current_row}, "Selesai")')
            cell_total.alignment = align_center
            cell_total.font = bold_data_font
            cell_total.border = thin_border
            if is_even: cell_total.fill = row_fill
            
            # Col J: Nilai Akhir
            # Formula: =IF(H5="Y","E",IF(I5=4,"A",IF(I5=3,"B+",IF(I5=2,"B",IF(I5=1,"D","E")))))
            cell_nilai = ws.cell(row=current_row, column=10, value=f'=IF(H{current_row}="Y","E",IF(I{current_row}=4,"A",IF(I{current_row}=3,"B+",IF(I{current_row}=2,"B",IF(I{current_row}=1,"D","E")))))')
            cell_nilai.alignment = align_center
            cell_nilai.font = bold_data_font
            cell_nilai.border = thin_border
            if is_even: cell_nilai.fill = row_fill
            
            # Col K: Status
            # Formula: =IF(J5="E","BELUM LULUS","LULUS")
            cell_status = ws.cell(row=current_row, column=11, value=f'=IF(J{current_row}="E","BELUM LULUS","LULUS")')
            cell_status.alignment = align_center
            cell_status.font = bold_data_font
            cell_status.border = thin_border
            if is_even: cell_status.fill = row_fill
            
            current_row += 1
            
        max_row = current_row - 1
        
        # Add Data Validation (Dropdown lists)
        # Dropdown Task 1-4: Selesai, Belum
        dv_task = DataValidation(type="list", formula1='"Selesai,Belum"', allow_blank=True)
        ws.add_data_validation(dv_task)
        dv_task.add(f"D5:G{max_row}")
        
        # Dropdown Warning Merah: Y, T
        dv_warning = DataValidation(type="list", formula1='"Y,T"', allow_blank=True)
        ws.add_data_validation(dv_warning)
        dv_warning.add(f"H5:H{max_row}")
        
        # Set Column Widths
        col_widths = {
            'A': 6,
            'B': 14,
            'C': 38,
            'D': 14,
            'E': 14,
            'F': 14,
            'G': 14,
            'H': 16,
            'I': 20,
            'J': 14,
            'K': 18
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

    # Sheet 2: Penilaian Kelas B
    ws_b = wb.create_sheet(title="Penilaian Kelas B")
    populate_class_sheet(ws_b, "Kelas B", "Sistem Penelitian Terintegrasi untuk JurnalMu", class_b_data)
    
    # Sheet 3: Penilaian Kelas G
    ws_g = wb.create_sheet(title="Penilaian Kelas G")
    populate_class_sheet(ws_g, "Kelas G", "Submission System Terintegrasi (OJS-based) untuk JurnalMu", class_g_data)
    
    # Ensure docs directory exists
    os.makedirs("docs", exist_ok=True)
    output_path = os.path.join("docs", "Penilaian_Mahasiswa_RPL_2026.xlsx")
    wb.save(output_path)
    print(f"Spreadsheet berhasil dibuat di: {output_path}")

if __name__ == "__main__":
    create_assessment_workbook()
